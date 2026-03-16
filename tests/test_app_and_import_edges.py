import io
import logging
import socket
from pathlib import Path

import openpyxl
import pytest
from flask import Flask
from werkzeug.datastructures import FileStorage

import app as app_module
from app import (
    create_app,
    ensure_runtime_schema_compatibility,
    find_free_port,
    open_browser,
    schedule_browser_open,
    seed_default_user,
)
from config import config
from database.db import db
from models.user import User
from services.auth_service import build_recovery_key_email_body
from services.import_service import ImportValidationError, import_finances_from_file
from services.logging_utils import configure_application_logging, request_id_context

GENERIC_RESPONSE_WARNING_MARKER = b'resposta generica'


def test_set_language_redirects_to_last_safe_page(client):
    with client.session_transaction() as session:
        session['language_redirect_target'] = '/about'

    response = client.get('/set_language/en', follow_redirects=False)

    assert response.status_code == 302
    assert response.headers['Location'].endswith('/about')


def test_set_language_preserves_last_safe_query_string(client):
    with client.session_transaction() as session:
        session['language_redirect_target'] = '/dashboard/2026/3?page=2'

    response = client.get('/set_language/en', follow_redirects=False)

    assert response.status_code == 302
    assert response.headers['Location'].endswith('/dashboard/2026/3?page=2')


def test_set_language_rejects_unsafe_session_redirect_target(client):
    with client.session_transaction() as session:
        session['language_redirect_target'] = 'http://evil.example/phish'

    response = client.get('/set_language/en', follow_redirects=False)

    assert response.status_code == 302
    assert response.headers['Location'].endswith('/')


def test_request_id_header_is_generated_when_missing(client):
    response = client.get('/health')

    assert response.status_code == 200
    assert response.headers['X-Request-ID']


def test_request_id_header_reuses_incoming_value(client):
    response = client.get('/health', headers={'X-Request-ID': 'request-123'})

    assert response.status_code == 200
    assert response.headers['X-Request-ID'] == 'request-123'


def test_configure_application_logging_writes_rotating_file(tmp_path):
    app = Flask('logging-test')
    app.config.update(
        LOG_LEVEL='INFO',
        SQLALCHEMY_LOG_LEVEL='WARNING',
        WERKZEUG_LOG_LEVEL='INFO',
        WAITRESS_LOG_LEVEL='INFO',
        LOG_FORMAT='%(levelname)s request_id=%(request_id)s %(message)s',
        LOG_TO_FILE=True,
        LOG_DIRECTORY=str(tmp_path),
        LOG_FILE_NAME='finora.log',
        LOG_MAX_BYTES=4096,
        LOG_BACKUP_COUNT=2,
    )
    app.logger.handlers.clear()

    configure_application_logging(app)
    token = request_id_context.set('req-file-123')
    try:
        app.logger.info('arquivo de log ativo')
    finally:
        request_id_context.reset(token)

    for handler in app.logger.handlers:
        handler.flush()
        handler.close()

    content = (Path(tmp_path) / 'finora.log').read_text(encoding='utf-8')
    assert 'arquivo de log ativo' in content
    assert 'request_id=req-file-123' in content


def test_configure_application_logging_keeps_sqlalchemy_quiet_by_default():
    app = Flask('logging-level-test')
    app.config.update(
        LOG_LEVEL='INFO',
        SQLALCHEMY_LOG_LEVEL='WARNING',
        WERKZEUG_LOG_LEVEL='INFO',
        WAITRESS_LOG_LEVEL='INFO',
    )

    configure_application_logging(app)

    assert logging.getLogger('sqlalchemy.engine').level == logging.WARNING


def test_health_endpoint_reports_degraded_when_db_fails(client, monkeypatch):
    def raise_db_down(*_args, **_kwargs):
        raise RuntimeError('db down')

    monkeypatch.setattr(app_module.db.session, 'execute', raise_db_down)
    response = client.get('/health')

    assert response.status_code == 503
    assert response.get_json() == {'status': 'degraded', 'database': 'error'}


def test_create_app_requires_secret_key_in_production(monkeypatch):
    original_secret = config['production'].SECRET_KEY
    config['production'].SECRET_KEY = None
    try:
        with pytest.raises(RuntimeError):
            create_app('production')
    finally:
        config['production'].SECRET_KEY = original_secret


def test_runtime_apps_are_isolated_from_local_database_during_tests():
    app = create_app('development')

    assert 'database/finora.db' not in app.config['SQLALCHEMY_DATABASE_URI']
    assert app.config['ENABLE_DEFAULT_USER_SEED'] is False
    assert app.config['ENABLE_RECURRING_SCHEDULER'] is False
    assert app.config['ENABLE_BACKUP_SCHEDULER'] is False


def test_seed_default_user_creates_user_when_enabled(app, monkeypatch):
    with app.app_context():
        db.create_all()
        app.config['TESTING'] = False
        app.config['ENABLE_DEFAULT_USER_SEED'] = True
        monkeypatch.setenv('DEFAULT_USER_NAME', 'Seed User')
        monkeypatch.setenv('DEFAULT_USER_USERNAME', 'seeduser')
        monkeypatch.setenv('DEFAULT_USER_EMAIL', 'seed@example.com')
        monkeypatch.setenv('DEFAULT_USER_PASSWORD', 'SeedPassword123')

        seed_default_user(app)

        seeded_user = User.query.filter_by(username='seeduser').first()
        assert seeded_user is not None
        assert seeded_user.email == 'seed@example.com'
        assert seeded_user.name == 'Seed User'


def test_forgot_password_page_does_not_render_privacy_warning(client):
    response = client.get('/forgot_password')

    assert response.status_code == 200
    assert GENERIC_RESPONSE_WARNING_MARKER not in response.data


def test_context_payloads_are_cached_per_locale(app, client, monkeypatch):
    call_counts = {
        'budget_categories': 0,
        'finance_catalog': 0,
        'payment_methods': 0,
    }

    def fake_budget_categories():
        call_counts['budget_categories'] += 1
        return ('Lazer',)

    def fake_finance_catalog():
        call_counts['finance_catalog'] += 1
        return {'Despesa': {'label': 'Despesa', 'categories': []}}

    def fake_payment_methods():
        call_counts['payment_methods'] += 1
        return [{'value': 'Dinheiro', 'label': 'Dinheiro'}]

    app.extensions.pop('finora_ui_payload_cache', None)
    monkeypatch.setattr(app_module, 'get_expense_budget_categories', fake_budget_categories)
    monkeypatch.setattr(app_module, 'build_finance_catalog_payload', fake_finance_catalog)
    monkeypatch.setattr(app_module, 'build_payment_method_payload', fake_payment_methods)

    with client.session_transaction() as session:
        session['lang'] = 'pt'

    first_response = client.get('/login')
    second_response = client.get('/login')

    assert first_response.status_code == 200
    assert second_response.status_code == 200
    assert call_counts == {
        'budget_categories': 1,
        'finance_catalog': 1,
        'payment_methods': 1,
    }


def test_context_payload_cache_translates_per_request_locale_without_rebuilding(app, monkeypatch):
    call_counts = {
        'budget_categories': 0,
        'finance_catalog': 0,
        'payment_methods': 0,
    }

    def fake_budget_categories():
        call_counts['budget_categories'] += 1
        return ('Utilities',)

    def fake_finance_catalog():
        call_counts['finance_catalog'] += 1
        return {'Expense': {'label': 'Catalog Label', 'categories': ['Utilities']}}

    def fake_payment_methods():
        call_counts['payment_methods'] += 1
        return [{'value': 'cash', 'label': 'Cash Label'}]

    def fake_gettext(value):
        from flask import session

        return f"{session.get('lang', 'pt')}::{value}"

    app.extensions.pop('finora_ui_payload_cache', None)
    monkeypatch.setattr(app_module, 'get_expense_budget_categories', fake_budget_categories)
    monkeypatch.setattr(app_module, 'build_finance_catalog_payload', fake_finance_catalog)
    monkeypatch.setattr(app_module, 'build_payment_method_payload', fake_payment_methods)
    monkeypatch.setattr(app_module, '_', fake_gettext)

    with app.test_request_context('/'):
        from flask import session

        session['lang'] = 'pt'
        payload_pt = app_module._get_cached_ui_payloads(app)

    with app.test_request_context('/'):
        from flask import session

        session['lang'] = 'en'
        payload_en = app_module._get_cached_ui_payloads(app)

    assert payload_pt['finance_catalog_payload']['Expense']['label'] == 'pt::Catalog Label'
    assert payload_en['finance_catalog_payload']['Expense']['label'] == 'en::Catalog Label'
    assert payload_pt['payment_method_options'][0]['label'] == 'pt::Cash Label'
    assert payload_en['payment_method_options'][0]['label'] == 'en::Cash Label'
    assert call_counts == {
        'budget_categories': 1,
        'finance_catalog': 1,
        'payment_methods': 1,
    }


def test_context_payload_cache_rebuilds_when_signature_changes(app, monkeypatch):
    call_count = {'finance_catalog': 0}
    signature_state = {'value': ('1.3.0', 1.0)}

    def fake_budget_categories():
        return ('Utilities',)

    def fake_finance_catalog():
        call_count['finance_catalog'] += 1
        return {'Expense': {'label': f"Catalog {call_count['finance_catalog']}", 'categories': []}}

    def fake_payment_methods():
        return [{'value': 'cash', 'label': 'Cash Label'}]

    app.extensions.pop('finora_ui_payload_cache', None)
    monkeypatch.setattr(app_module, 'get_expense_budget_categories', fake_budget_categories)
    monkeypatch.setattr(app_module, 'build_finance_catalog_payload', fake_finance_catalog)
    monkeypatch.setattr(app_module, 'build_payment_method_payload', fake_payment_methods)
    monkeypatch.setattr(
        app_module,
        '_build_ui_payload_cache_signature',
        lambda _app: signature_state['value'],
    )
    monkeypatch.setattr(app_module, 'UI_PAYLOAD_CACHE_REFRESH_INTERVAL_SECONDS', 0.0)

    with app.test_request_context('/'):
        first_payload = app_module._get_cached_ui_payloads(app)

    signature_state['value'] = ('1.3.1', 2.0)

    with app.test_request_context('/'):
        second_payload = app_module._get_cached_ui_payloads(app)

    assert first_payload['finance_catalog_payload']['Expense']['label'] == 'Catalog 1'
    assert second_payload['finance_catalog_payload']['Expense']['label'] == 'Catalog 2'
    assert call_count['finance_catalog'] == 2


def test_auth_templates_use_premium_cards(client):
    login_html = client.get('/login').get_data(as_text=True)
    register_html = client.get('/register').get_data(as_text=True)

    assert 'card card-premium auth-card-premium border-0' in login_html
    assert 'card card-premium auth-card-premium border-0' in register_html
    assert 'data-loading-text="' in login_html
    assert 'data-loading-text="' in register_html


@pytest.mark.parametrize(
    ('lang_code', 'expected_text'),
    [
        ('en', 'We recommend a strong password.'),
        ('es', 'Recomendamos una contraseña fuerte.'),
    ],
)
def test_register_page_translates_password_guidance(client, lang_code, expected_text):
    with client.session_transaction() as session:
        session['lang'] = lang_code

    response = client.get('/register')

    assert response.status_code == 200
    assert expected_text in response.get_data(as_text=True)


def test_welcome_page_translates_marketing_copy_in_spanish(client):
    with client.session_transaction() as session:
        session['lang'] = 'es'

    response = client.get('/')
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert 'Control financiero local e inteligente' in html
    assert 'Sigue cada movimiento con estado y vencimiento.' in html
    assert 'Confiabilidad' in html


def test_about_page_translates_update_section_in_english(client):
    with client.session_transaction() as session:
        session['lang'] = 'en'

    response = client.get('/about')
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert 'System Updates' in html
    assert 'Installed version' in html
    assert 'Check for updates' in html
    assert 'The bundled local manifest is active for status and safe testing.' in html


def test_register_lookup_endpoints_are_non_disclosive(client):
    username_response = client.post('/check_username', json={'username': 'someone'})
    email_response = client.post('/check_email', json={'email': 'someone@example.com'})

    assert username_response.status_code == 200
    assert email_response.status_code == 200
    assert username_response.get_json()['verified'] is False
    assert email_response.get_json()['verified'] is False


def test_dashboard_period_form_uses_loading_submit_feedback(client, app):
    with app.app_context():
        user = User(username='periodloading', email='periodloading@example.com', name='Period Loading')
        user.set_password('Strong123')
        db.session.add(user)
        db.session.commit()

    client.post('/login', data={'identifier': 'periodloading', 'password': 'Strong123'}, follow_redirects=True)
    html = client.get('/dashboard/2026/3').get_data(as_text=True)

    assert 'data-loading-text="' in html
    assert 'data-edit-action-template="' in html


def test_recovery_key_email_body_is_translated_in_english(app):
    with app.app_context():
        user = User(username='mailuser', email='mailuser@example.com', name='Mail User')
        user.recovery_key_version = 2

        with app.test_request_context('/'):
            from flask import session

            session['lang'] = 'en'
            body = build_recovery_key_email_body(user, 'ABCD1234EFGH5678', 'resend')

    assert 'Hello, Mail User!' in body
    assert 'As requested, we have resent your recovery key.' in body
    assert 'Recovery key: ABCD1234EFGH5678' in body
    assert 'Key version: 2' in body


def test_profile_page_translates_hub_labels_in_english(app, client):
    with app.app_context():
        user = User(username='profileen', email='profileen@example.com', name='Profile EN')
        user.set_password('Strong123')
        db.session.add(user)
        db.session.commit()

    with client.session_transaction() as session:
        session['lang'] = 'en'

    client.post('/login', data={'identifier': 'profileen', 'password': 'Strong123'}, follow_redirects=True)

    with client.session_transaction() as session:
        session['lang'] = 'en'

    response = client.get('/profile')
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert 'My Backups' in html
    assert 'Active sessions' in html
    assert 'Recent activity' in html
    assert 'System Status' in html
    assert 'Send Suggestion' in html
    assert 'Report Issue' in html
    assert 'Updates' in html
    assert 'Status: Ready' in html
    assert 'subject=Finora%20-%20Suggestion' in html


def test_profile_page_translates_hub_labels_in_spanish(app, client):
    with app.app_context():
        user = User(username='profilees', email='profilees@example.com', name='Profile ES')
        user.set_password('Strong123')
        db.session.add(user)
        db.session.commit()

    with client.session_transaction() as session:
        session['lang'] = 'es'

    client.post('/login', data={'identifier': 'profilees', 'password': 'Strong123'}, follow_redirects=True)

    with client.session_transaction() as session:
        session['lang'] = 'es'

    response = client.get('/profile')
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert 'Mis copias' in html
    assert 'Sesiones activas' in html
    assert 'Actividad reciente' in html
    assert 'Estado del sistema' in html
    assert 'Enviar sugerencia' in html
    assert 'Reportar error' in html
    assert 'Actualizaciones' in html
    assert 'Estado: Listo' in html
    assert 'subject=Finora%20-%20Sugerencia' in html


def test_ensure_runtime_schema_compatibility_is_noop_under_testing(app):
    with app.app_context():
        ensure_runtime_schema_compatibility(app)


def test_find_free_port_returns_available_port():
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as temp_socket:
        temp_socket.bind(('127.0.0.1', 0))
        occupied_port = temp_socket.getsockname()[1]
        free_port = find_free_port(occupied_port)

        assert isinstance(free_port, int)
        assert free_port >= occupied_port
        assert free_port != occupied_port

    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as test_socket:
        test_socket.bind(('127.0.0.1', free_port))


def test_open_browser_delegates_to_webbrowser(monkeypatch):
    opened = {}
    monkeypatch.setattr(
        'webbrowser.open',
        lambda url, new=0, autoraise=True: opened.__setitem__(
            'call',
            (url, new, autoraise),
        ),
    )

    open_browser(5050)

    assert opened['call'] == ('http://127.0.0.1:5050/', 0, True)


def test_schedule_browser_open_starts_timer_once(monkeypatch):
    calls = {}

    class FakeTimer:
        def __init__(self, delay, callback):
            calls['delay'] = delay
            calls['callback'] = callback

        def start(self):
            calls['started'] = True

    monkeypatch.setattr(app_module, 'Timer', FakeTimer)
    monkeypatch.setattr(app_module, '_acquire_browser_launch_guard', lambda _port: True)
    monkeypatch.setenv('FINORA_AUTO_OPEN_BROWSER', '1')

    assert schedule_browser_open(5050) is True
    assert calls == {
        'delay': 1.5,
        'callback': calls['callback'],
        'started': True,
    }


def test_schedule_browser_open_skips_when_guard_is_active(monkeypatch):
    monkeypatch.setattr(app_module, '_acquire_browser_launch_guard', lambda _port: False)
    monkeypatch.setenv('FINORA_AUTO_OPEN_BROWSER', '1')

    assert schedule_browser_open(5050) is False


def test_schedule_browser_open_can_be_disabled(monkeypatch):
    monkeypatch.setenv('FINORA_AUTO_OPEN_BROWSER', '0')
    monkeypatch.setattr(
        app_module,
        'Timer',
        lambda *_args, **_kwargs: pytest.fail('Timer should not be created'),
    )

    assert schedule_browser_open(5050) is False


def test_import_finances_from_xlsx_supports_aliases():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['Descrição', 'Valor', 'Categoria', 'Tipo', 'Situação', 'Data'])
    sheet.append(['Salário', 5000, 'Salário', 'Receita', 'Pago', '2026-03-10'])

    stream = io.BytesIO()
    try:
        workbook.save(stream)
    finally:
        workbook.close()
    stream.seek(0)

    uploaded = FileStorage(stream=stream, filename='finances.xlsx')
    result = import_finances_from_file(uploaded, user_id=1)

    assert result.imported_rows == 1
    assert result.entries[0].description == 'Salário'
    assert result.entries[0].category == 'Trabalho'
    assert result.entries[0].subcategory == 'Salário'
    assert result.entries[0].type == 'Receita'


def test_import_finances_from_xlsx_supports_english_aliases():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['description', 'value', 'category', 'type', 'status', 'date', 'payment method'])
    sheet.append(['Salary', 5000, 'Salário', 'Income', 'Paid', '2026-03-10', 'credit card'])

    stream = io.BytesIO()
    try:
        workbook.save(stream)
    finally:
        workbook.close()
    stream.seek(0)

    uploaded = FileStorage(stream=stream, filename='finances.xlsx')
    result = import_finances_from_file(uploaded, user_id=1)

    assert result.imported_rows == 1
    assert result.entries[0].description == 'Salary'
    assert result.entries[0].category == 'Trabalho'
    assert result.entries[0].subcategory == 'Salário'
    assert result.entries[0].type == 'Receita'
    assert result.entries[0].status == 'Pago'
    assert result.entries[0].payment_method == 'Cartão de Crédito'


def test_import_finances_rejects_txt_extension():
    uploaded = FileStorage(stream=io.BytesIO(b'test'), filename='finances.txt')

    with pytest.raises(ImportValidationError):
        import_finances_from_file(uploaded, user_id=1)


def test_import_finances_rejects_pdf_extension():
    uploaded = FileStorage(stream=io.BytesIO(b'test'), filename='finances.pdf')

    with pytest.raises(ImportValidationError):
        import_finances_from_file(uploaded, user_id=1)


def test_import_finances_rejects_doc_extension():
    uploaded = FileStorage(stream=io.BytesIO(b'test'), filename='finances.doc')

    with pytest.raises(ImportValidationError):
        import_finances_from_file(uploaded, user_id=1)


def test_import_finances_rejects_file_above_row_limit():
    csv_content = (
        'descricao,valor,categoria,tipo,status,data\n'
        'Item 1,10,Lazer,Despesa,Pago,2026-03-10\n'
        'Item 2,20,Lazer,Despesa,Pago,2026-03-11\n'
    )
    uploaded = FileStorage(stream=io.BytesIO(csv_content.encode('utf-8')), filename='finances.csv')

    with pytest.raises(ImportValidationError, match='limite de 1 linha permitida'):
        import_finances_from_file(uploaded, user_id=1, max_rows=1)


def test_import_finances_rejects_zero_value_entries():
    # The in-memory CSV payload has a single row with value 0, which is invalid
    # because imports only accept amounts greater than zero.
    csv_content = (
        'descricao,valor,categoria,tipo,status,data\n'
        'Item inválido,0,Lazer,Despesa,Pago,2026-03-10\n'
    )
    uploaded = FileStorage(stream=io.BytesIO(csv_content.encode('utf-8')), filename='finances.csv')

    with pytest.raises(ImportValidationError, match='Valor deve ser maior que zero'):
        import_finances_from_file(uploaded, user_id=1)


def test_import_finances_rejects_negative_values():
    csv_content = (
        'descricao,valor,categoria,tipo,status,data\n'
        'Item inválido,-10,Lazer,Despesa,Pago,2026-03-10\n'
    )
    uploaded = FileStorage(stream=io.BytesIO(csv_content.encode('utf-8')), filename='finances.csv')

    with pytest.raises(ImportValidationError, match='Valor deve ser maior que zero'):
        import_finances_from_file(uploaded, user_id=1)


def test_import_finances_accepts_decimal_values():
    csv_content = (
        'descricao,valor,categoria,tipo,status,data\n'
        'Freela,10.75,Salário,Receita,Pago,2026-03-10\n'
    )
    uploaded = FileStorage(stream=io.BytesIO(csv_content.encode('utf-8')), filename='finances.csv')

    result = import_finances_from_file(uploaded, user_id=1)

    assert result.imported_rows == 1
    assert result.entries[0].value == 10.75


def test_import_finances_from_xlsx_supports_mixed_language_aliases():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['Descrição', 'value', 'Categoria', 'type', 'Situação', 'Data'])
    sheet.append(['Projeto', 800, 'Salário', 'Income', 'Paid', '2026-03-10'])

    stream = io.BytesIO()
    try:
        workbook.save(stream)
    finally:
        workbook.close()
    stream.seek(0)

    uploaded = FileStorage(stream=stream, filename='finances.xlsx')
    result = import_finances_from_file(uploaded, user_id=1)

    assert result.imported_rows == 1
    assert result.entries[0].description == 'Projeto'
    assert result.entries[0].category == 'Trabalho'
    assert result.entries[0].subcategory == 'Salário'
    assert result.entries[0].type == 'Receita'
    assert result.entries[0].status == 'Pago'
