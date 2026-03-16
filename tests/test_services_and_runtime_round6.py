import io
import os
from datetime import date, timedelta

import openpyxl
import pytest
from werkzeug.datastructures import FileStorage

from app import create_app, ensure_runtime_schema_compatibility, find_free_port, seed_default_user
from database.db import db
from models.backup import BackupRecord, BackupSchedule
from models.recurring import RecurringEntry
from models.user import User
from routes import backup as backup_module
from services import backup_service, import_service, recurring_service
from services.import_service import ImportValidationError


def _create_user(app, username, email, password='Password123'):
    with app.app_context():
        user = User(username=username, email=email, name=username.title())
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        return user.id


def _login(client, username, password='Password123'):
    return client.post(
        '/login',
        data={'identifier': username, 'password': password},
        follow_redirects=True,
    )


def test_backup_download_rejects_non_sqlite_engine(auth_client, monkeypatch):
    monkeypatch.setattr(
        backup_module,
        'create_backup_for_user',
        lambda *_args, **_kwargs: (_ for _ in ()).throw(ValueError('Backup por arquivo está disponível apenas para SQLite.')),
    )

    response = auth_client.get('/backup/download', follow_redirects=True)

    assert response.status_code == 200
    assert b'apenas para SQLite' in response.data


def test_backup_download_rejects_missing_database_file(auth_client, monkeypatch):
    monkeypatch.setattr(
        backup_module,
        'create_backup_for_user',
        lambda *_args, **_kwargs: (_ for _ in ()).throw(ValueError('Banco de dados não encontrado para backup.')),
    )

    response = auth_client.get('/backup/download', follow_redirects=True)

    assert response.status_code == 200
    assert b'Banco de dados n\xc3\xa3o encontrado para backup' in response.data


def test_backup_schedule_update_persists_configuration(auth_client, app):
    response = auth_client.post(
        '/backup/schedule',
        data={
            'enabled': 'on',
            'frequency': 'Semanal',
            'times_per_period': '2',
            'day_of_week': '4',
            'run_hour': '9',
            'run_minute': '30',
            'retention_count': '6',
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b'Rotina de backup atualizada com sucesso' in response.data

    with app.app_context():
        user = User.query.filter_by(username='testuser').first()
        schedule = BackupSchedule.query.filter_by(user_id=user.id).first()
        assert schedule is not None
        assert schedule.enabled is True
        assert schedule.frequency == 'Semanal'
        assert schedule.times_per_period == 2
        assert schedule.day_of_week == 4
        assert schedule.retention_count == 6
        assert schedule.next_run_at is not None


def test_apply_backup_schedule_update_rejects_invalid_frequency(app):
    with app.app_context():
        user = User(username='backupinvalid', email='backupinvalid@example.com', name='Backup Invalid')
        user.set_password('Password123')
        db.session.add(user)
        db.session.commit()

        error_code = backup_service.apply_backup_schedule_update(
            user,
            form={
                'enabled': 'on',
                'frequency': 'Anual',
                'times_per_period': '1',
                'run_hour': '3',
                'run_minute': '0',
                'retention_count': '5',
            },
            default_retention_count=20,
        )

        assert error_code == 'invalid_frequency'


def test_run_backup_maintenance_processes_due_schedule(tmp_path):
    db_path = tmp_path / 'backup-maintenance.db'
    app = create_app('development')
    app.config.update(
        TESTING=False,
        WTF_CSRF_ENABLED=False,
        SQLALCHEMY_DATABASE_URI=f"sqlite:///{db_path.as_posix()}",
        BACKUP_STORAGE_DIR=str(tmp_path / 'backups'),
        ENABLE_RECURRING_SCHEDULER=False,
    )

    with app.app_context():
        db.drop_all()
        db.create_all()
        user = User(username='backupauto', email='backupauto@example.com', name='Backup Auto')
        user.set_password('Password123')
        db.session.add(user)
        db.session.commit()

        schedule = BackupSchedule(
            user_id=user.id,
            enabled=True,
            frequency='Diário',
            times_per_period=1,
            run_hour=1,
            run_minute=0,
            retention_count=1,
            next_run_at=backup_service.utcnow_naive() - timedelta(minutes=5),
        )
        db.session.add(schedule)
        db.session.commit()

        result = backup_service.run_backup_maintenance(app)

        assert result['processed_backups'] == 1
        assert result['affected_users'] == 1
        assert BackupRecord.query.filter_by(user_id=user.id).count() == 1
        refreshed_schedule = db.session.get(BackupSchedule, schedule.id)
        assert refreshed_schedule.last_run_at is not None
        assert refreshed_schedule.next_run_at is not None

        db.session.remove()
        db.drop_all()
        db.engine.dispose()


def test_delete_saved_backup_route_removes_record_and_file(tmp_path):
    db_path = tmp_path / 'backup-delete.db'
    app = create_app('development')
    app.config.update(
        TESTING=True,
        WTF_CSRF_ENABLED=False,
        SQLALCHEMY_DATABASE_URI=f"sqlite:///{db_path.as_posix()}",
        BACKUP_STORAGE_DIR=str(tmp_path / 'backups'),
        ENABLE_RECURRING_SCHEDULER=False,
    )

    with app.app_context():
        db.drop_all()
        db.create_all()
        user = User(username='backupdelete', email='backupdelete@example.com', name='Backup Delete')
        user.set_password('Password123')
        db.session.add(user)
        db.session.commit()
        user_id = user.id

        backup_dir = tmp_path / 'backups' / f'user_{user_id}'
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_file = backup_dir / 'manual.zip'
        backup_file.write_bytes(b'PKdemo')

        record = BackupRecord(
            user_id=user_id,
            trigger_source='Manual',
            status='Concluido',
            file_name='manual.zip',
            storage_path=str(backup_file),
            file_size_bytes=backup_file.stat().st_size,
            checksum='checksum',
        )
        db.session.add(record)
        db.session.commit()
        record_id = record.id

    client = app.test_client()
    _login(client, 'backupdelete')
    response = client.post(f'/backup/history/{record_id}/delete', follow_redirects=True)

    assert response.status_code == 200
    assert b'Backup removido com sucesso' in response.data

    with app.app_context():
        assert db.session.get(BackupRecord, record_id) is None
        assert not os.path.exists(backup_file)
        db.session.remove()
        db.drop_all()
        db.engine.dispose()


def test_run_backup_maintenance_skips_when_schema_is_incomplete(app, monkeypatch):
    monkeypatch.setattr(backup_service, 'backup_schema_is_ready', lambda: False)

    result = backup_service.run_backup_maintenance(app)

    assert result == {'processed_backups': 0, 'affected_users': 0, 'skipped': True}


def test_start_backup_scheduler_returns_none_in_testing(app):
    assert backup_service.start_backup_scheduler(app) is None


def test_seed_default_user_skips_when_password_missing(app, monkeypatch):
    with app.app_context():
        db.create_all()
        app.config['TESTING'] = False
        app.config['ENABLE_DEFAULT_USER_SEED'] = True
        monkeypatch.delenv('DEFAULT_USER_PASSWORD', raising=False)

        seed_default_user(app)

        assert User.query.count() == 0


def test_ensure_runtime_schema_compatibility_skips_non_sqlite(app):
    app.config['TESTING'] = False
    app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://example'

    ensure_runtime_schema_compatibility(app)


def test_find_free_port_returns_start_port_when_range_is_exhausted():
    assert find_free_port(65535) == 65535


def test_import_service_helpers_cover_csv_xlsx_and_normalizers():
    empty_upload = FileStorage(stream=io.BytesIO(b''), filename='finances.csv')
    with pytest.raises(ImportValidationError):
        import_service._validate_upload(empty_upload, max_file_size=1024)

    oversized_upload = FileStorage(stream=io.BytesIO(b'a' * 5), filename='finances.csv')
    with pytest.raises(ImportValidationError):
        import_service._validate_upload(oversized_upload, max_file_size=1)

    csv_upload = FileStorage(
        stream=io.BytesIO('descricao;valor\n\nCafe;10\n'.encode('utf-8')),
        filename='finances.csv',
    )
    rows = import_service._read_csv_rows(csv_upload)
    assert rows == [(2, {'descricao': 'Cafe', 'valor': '10'})]

    invalid_csv_upload = FileStorage(stream=io.BytesIO(b'\n'), filename='finances.csv')
    with pytest.raises(ImportValidationError):
        import_service._read_csv_rows(invalid_csv_upload)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append([None, None])
    xlsx_stream = io.BytesIO()
    workbook.save(xlsx_stream)
    xlsx_stream.seek(0)
    with pytest.raises(ImportValidationError):
        import_service._read_xlsx_rows(FileStorage(stream=xlsx_stream, filename='blank-header.xlsx'))

    empty_workbook = openpyxl.Workbook()
    empty_sheet = empty_workbook.active
    empty_sheet.delete_rows(1, empty_sheet.max_row)
    empty_stream = io.BytesIO()
    empty_workbook.save(empty_stream)
    empty_stream.seek(0)
    with pytest.raises(ImportValidationError):
        import_service._read_xlsx_rows(FileStorage(stream=empty_stream, filename='empty.xlsx'))

    canonical = import_service._to_canonical_fields({'Data Vencimento': '2026-03-10', 'Tipo': 'Receita'})
    assert canonical == {'due_date': '2026-03-10', 'type': 'Receita'}

    assert import_service._normalize_type(None) == 'Despesa'
    assert import_service._normalize_status(None) == 'Pendente'
    assert import_service._normalize_type('receita') == 'Receita'
    assert import_service._normalize_status('atrasado') == 'Atrasado'

    assert float(import_service._parse_money('1.234,56')) == pytest.approx(1234.56)
    with pytest.raises(ImportValidationError):
        import_service._parse_money(None)
    with pytest.raises(ImportValidationError):
        import_service._parse_money('abc')

    assert import_service._parse_date('10-03-2026', required=False) == date(2026, 3, 10)
    with pytest.raises(ImportValidationError):
        import_service._parse_date('', required=True)
    with pytest.raises(ImportValidationError):
        import_service._parse_date(10**12, required=False)
    with pytest.raises(ImportValidationError):
        import_service._parse_date('2026/03/10', required=False)

    assert import_service._is_empty_row([None, '   ']) is True
    assert import_service._is_empty_row([None, 'valor']) is False


def test_recurring_service_handles_invalid_frequency_and_end_date(app):
    with app.app_context():
        user = User(username='recurend', email='recurend@example.com', name='Recur End')
        user.set_password('Password123')
        db.session.add(user)
        db.session.commit()
        recurring = RecurringEntry(
            description='Invalid Recurring',
            value=10.0,
            category='Lazer',
            type='Despesa',
            frequency='Invalida',
            start_date=date.today() - timedelta(days=3),
            next_run_date=date.today() - timedelta(days=1),
            end_date=date.today() - timedelta(days=2),
            user_id=user.id,
        )
        db.session.add(recurring)
        db.session.commit()

        processed = recurring_service.process_recurring_entries(user.id, commit=False)

        assert processed == 0
        assert recurring.active is False
        assert recurring_service._advance_next_run_date(recurring) is False


def test_process_all_recurring_entries_skips_failed_users(monkeypatch):
    class FakeQuery:
        def filter(self, *_args, **_kwargs):
            return self

        def distinct(self):
            return self

        def all(self):
            return [(1,), (2,)]

    calls = {'rollback': 0}

    monkeypatch.setattr(recurring_service.db.session, 'query', lambda *_args, **_kwargs: FakeQuery())
    monkeypatch.setattr(recurring_service.db.session, 'rollback', lambda: calls.__setitem__('rollback', calls['rollback'] + 1))

    def fake_process(user_id):
        if user_id == 1:
            raise RuntimeError('boom')
        return 2

    monkeypatch.setattr(recurring_service, 'process_recurring_entries', fake_process)

    result = recurring_service.process_all_recurring_entries()

    assert result == {'processed_entries': 2, 'affected_users': 1}
    assert calls['rollback'] == 1
