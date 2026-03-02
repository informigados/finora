from __future__ import annotations

import csv
import io
import os
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage

from models.finance import Finance

MAX_IMPORT_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_IMPORT_ROWS = 20000
SUPPORTED_IMPORT_EXTENSIONS = {".csv", ".xlsx"}

TYPE_MAP = {
    "receita": "Receita",
    "despesa": "Despesa",
}

STATUS_MAP = {
    "pago": "Pago",
    "pendente": "Pendente",
    "atrasado": "Atrasado",
}

COLUMN_ALIASES = {
    "description": "description",
    "descricao": "description",
    "desc": "description",
    "value": "value",
    "valor": "value",
    "category": "category",
    "categoria": "category",
    "type": "type",
    "tipo": "type",
    "status": "status",
    "situacao": "status",
    "due_date": "due_date",
    "due date": "due_date",
    "datavencimento": "due_date",
    "vencimento": "due_date",
    "data": "due_date",
    "payment_date": "payment_date",
    "payment date": "payment_date",
    "pagamento": "payment_date",
    "observations": "observations",
    "observacoes": "observations",
}


class ImportValidationError(ValueError):
    pass


@dataclass
class ImportResult:
    entries: list[Finance]
    imported_rows: int
    skipped_rows: int
    errors: list[str]


def import_finances_from_file(
    uploaded_file: FileStorage,
    user_id: int,
    max_rows: int = MAX_IMPORT_ROWS,
    max_file_size: int = MAX_IMPORT_FILE_SIZE_BYTES,
) -> ImportResult:
    extension = _validate_upload(uploaded_file, max_file_size)

    if extension == ".csv":
        rows = _read_csv_rows(uploaded_file)
    elif extension == ".xlsx":
        rows = _read_xlsx_rows(uploaded_file)
    else:
        raise ImportValidationError("Formato de arquivo não suportado.")

    entries: list[Finance] = []
    errors: list[str] = []
    total_rows = 0

    for row_number, raw_row in rows:
        total_rows += 1
        if total_rows > max_rows:
            raise ImportValidationError(
                f"O arquivo excede o limite de {max_rows} linhas permitidas para importação."
            )

        try:
            entry = _build_entry_from_row(raw_row, user_id)
            entries.append(entry)
        except ImportValidationError as exc:
            errors.append(f"Linha {row_number}: {exc}")

    if not entries and errors:
        raise ImportValidationError(
            "Nenhum lançamento válido foi encontrado no arquivo enviado."
        )

    return ImportResult(
        entries=entries,
        imported_rows=len(entries),
        skipped_rows=len(errors),
        errors=errors,
    )


def _validate_upload(uploaded_file: FileStorage, max_file_size: int) -> str:
    filename = (uploaded_file.filename or "").strip()
    if not filename:
        raise ImportValidationError("Selecione um arquivo para importar.")

    extension = os.path.splitext(filename)[1].lower()
    if extension not in SUPPORTED_IMPORT_EXTENSIONS:
        readable_extensions = ", ".join(sorted(SUPPORTED_IMPORT_EXTENSIONS))
        raise ImportValidationError(
            f"Formato inválido. Utilize apenas: {readable_extensions}."
        )

    uploaded_file.stream.seek(0, os.SEEK_END)
    file_size = uploaded_file.stream.tell()
    uploaded_file.stream.seek(0)

    if file_size <= 0:
        raise ImportValidationError("O arquivo enviado está vazio.")
    if file_size > max_file_size:
        raise ImportValidationError(
            f"O arquivo excede o limite de {max_file_size // (1024 * 1024)} MB."
        )

    return extension


def _read_csv_rows(uploaded_file: FileStorage) -> list[tuple[int, dict[str, Any]]]:
    raw_bytes = uploaded_file.stream.read()
    uploaded_file.stream.seek(0)

    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if text is None:
        raise ImportValidationError(
            "Não foi possível decodificar o arquivo CSV. Use UTF-8 ou Windows-1252."
        )

    sample = text[:2048]
    delimiter = ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;|\t").delimiter
    except csv.Error:
        pass

    stream = io.StringIO(text, newline=None)
    reader = csv.DictReader(stream, delimiter=delimiter)
    if not reader.fieldnames:
        raise ImportValidationError("Cabeçalho do CSV não foi identificado.")

    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, row in enumerate(reader, start=2):
        if _is_empty_row(row.values()):
            continue
        rows.append((row_number, dict(row)))
    return rows


def _read_xlsx_rows(uploaded_file: FileStorage) -> list[tuple[int, dict[str, Any]]]:
    uploaded_file.stream.seek(0)
    workbook = load_workbook(uploaded_file.stream, read_only=True, data_only=True)

    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header = next(iterator, None)
        if not header:
            raise ImportValidationError("A planilha está vazia.")

        normalized_headers = [str(col).strip() if col is not None else "" for col in header]
        if all(not col for col in normalized_headers):
            raise ImportValidationError("Não foi possível identificar o cabeçalho da planilha.")

        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, values in enumerate(iterator, start=2):
            row_dict = {
                normalized_headers[idx]: values[idx] if idx < len(values) else None
                for idx in range(len(normalized_headers))
            }
            if _is_empty_row(row_dict.values()):
                continue
            rows.append((row_number, row_dict))
        return rows
    finally:
        workbook.close()


def _build_entry_from_row(raw_row: dict[str, Any], user_id: int) -> Finance:
    canonical = _to_canonical_fields(raw_row)

    description = str(canonical.get("description") or "Lançamento importado").strip()
    category = str(canonical.get("category") or "Geral").strip()

    amount = _parse_money(canonical.get("value"))
    if amount <= 0:
        raise ImportValidationError("Valor deve ser maior que zero.")

    entry_type = _normalize_type(canonical.get("type"))
    status = _normalize_status(canonical.get("status"))
    due_date = _parse_date(canonical.get("due_date"), required=False) or date.today()
    payment_date = _parse_date(canonical.get("payment_date"), required=False)
    observations = canonical.get("observations")
    observations_str = str(observations).strip() if observations is not None else None

    return Finance(
        description=description[:100],
        value=float(amount),
        category=category[:50],
        type=entry_type,
        status=status,
        due_date=due_date,
        payment_date=payment_date,
        observations=observations_str,
        user_id=user_id,
    )


def _to_canonical_fields(raw_row: dict[str, Any]) -> dict[str, Any]:
    canonical: dict[str, Any] = {}
    for key, value in raw_row.items():
        normalized_key = _normalize_text(str(key or ""))
        mapped_key = COLUMN_ALIASES.get(normalized_key)
        if not mapped_key and normalized_key.replace(" ", "") in COLUMN_ALIASES:
            mapped_key = COLUMN_ALIASES[normalized_key.replace(" ", "")]
        if mapped_key and mapped_key not in canonical:
            canonical[mapped_key] = value
    return canonical


def _normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return normalized.strip().lower().replace("_", " ")


def _normalize_type(raw_value: Any) -> str:
    if raw_value is None:
        return "Despesa"
    normalized = _normalize_text(str(raw_value)).replace(" ", "")
    return TYPE_MAP.get(normalized, "Despesa")


def _normalize_status(raw_value: Any) -> str:
    if raw_value is None:
        return "Pendente"
    normalized = _normalize_text(str(raw_value)).replace(" ", "")
    return STATUS_MAP.get(normalized, "Pendente")


def _parse_money(raw_value: Any) -> Decimal:
    if raw_value is None or str(raw_value).strip() == "":
        raise ImportValidationError("Valor ausente.")

    if isinstance(raw_value, (int, float, Decimal)):
        return Decimal(str(raw_value))

    text = str(raw_value).strip()
    text = text.replace("R$", "").replace(" ", "")

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ImportValidationError("Valor inválido.") from exc


def _parse_date(raw_value: Any, required: bool) -> date | None:
    if raw_value is None or str(raw_value).strip() == "":
        if required:
            raise ImportValidationError("Data obrigatória ausente.")
        return None

    if isinstance(raw_value, datetime):
        return raw_value.date()
    if isinstance(raw_value, date):
        return raw_value

    if isinstance(raw_value, (int, float)):
        excel_origin = date(1899, 12, 30)
        try:
            return excel_origin + timedelta(days=int(raw_value))
        except OverflowError as exc:
            raise ImportValidationError("Data inválida.") from exc

    text = str(raw_value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise ImportValidationError("Formato de data inválido.")


def _is_empty_row(values: Any) -> bool:
    for value in values:
        if value is None:
            continue
        if str(value).strip():
            return False
    return True
