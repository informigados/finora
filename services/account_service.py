from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from flask import has_request_context
from flask_babel import gettext as _
from sqlalchemy import case, func
from werkzeug.datastructures import FileStorage

from database.db import db
from models.account import (
    AccountTransfer,
    BankImportProfile,
    BankTransaction,
    FinancialAccount,
)
from models.finance import Finance
from models.time_utils import utcnow_naive


ALLOWED_ACCOUNT_TYPES = frozenset(
    {'checking', 'savings', 'wallet', 'credit_card', 'investment', 'other'}
)
ACCOUNT_COLOR_PALETTE = {
    '#2563EB': 'Azul',
    '#059669': 'Verde',
    '#7C3AED': 'Violeta',
    '#D97706': 'Âmbar',
    '#E11D48': 'Rosa',
    '#475569': 'Grafite',
}
ALLOWED_PROFILE_FILE_TYPES = {'csv', 'xlsx'}
MAX_STATEMENT_ROWS = 20000
MAX_STATEMENT_SIZE = 10 * 1024 * 1024


def _localized(message: str) -> str:
    """Translate labels in requests while keeping service calls context-safe."""
    return _(message) if has_request_context() else message


class AccountValidationError(ValueError):
    pass


@dataclass
class StatementImportResult:
    imported_rows: int
    duplicate_rows: int
    transactions: list[BankTransaction]


def get_account_type_options() -> list[tuple[str, str]]:
    return [
        ('checking', _localized('Conta corrente')),
        ('savings', _localized('Conta poupança')),
        ('wallet', _localized('Carteira')),
        ('credit_card', _localized('Cartão de crédito')),
        ('investment', _localized('Investimentos')),
        ('other', _localized('Outra conta')),
    ]


def get_account_type_label(account_type: str) -> str:
    return dict(get_account_type_options()).get(account_type, _localized('Outra conta'))


def get_account_color_options() -> list[tuple[str, str]]:
    return [
        ('#2563EB', _('Azul')),
        ('#059669', _('Verde')),
        ('#7C3AED', _('Violeta')),
        ('#D97706', _('Âmbar')),
        ('#E11D48', _('Rosa')),
        ('#475569', _('Grafite')),
    ]


def parse_decimal(value: Any) -> Decimal:
    if value is None or str(value).strip() == '':
        raise AccountValidationError(_('Informe um valor válido.'))
    if isinstance(value, (int, float, Decimal)):
        parsed = Decimal(str(value))
    else:
        normalized = str(value).strip().replace('R$', '').replace(' ', '')
        negative_parentheses = normalized.startswith('(') and normalized.endswith(')')
        normalized = normalized.strip('()')
        if ',' in normalized and '.' in normalized:
            if normalized.rfind(',') > normalized.rfind('.'):
                normalized = normalized.replace('.', '').replace(',', '.')
            else:
                normalized = normalized.replace(',', '')
        elif ',' in normalized:
            normalized = normalized.replace('.', '').replace(',', '.')
        try:
            parsed = Decimal(normalized)
        except InvalidOperation as exc:
            raise AccountValidationError(_('Informe um valor válido.')) from exc
        if negative_parentheses:
            parsed = -parsed
    return parsed.quantize(Decimal('0.01'))


def validate_account_payload(data) -> dict[str, Any]:
    name = (data.get('name') or '').strip()
    account_type = (data.get('account_type') or '').strip()
    institution = (data.get('institution') or '').strip()
    last_four = re.sub(r'\D', '', data.get('last_four') or '')[-4:]
    color = (data.get('color') or '#2563EB').strip().upper()
    initial_balance = parse_decimal(data.get('initial_balance') or '0')

    if not name or len(name) > 80:
        raise AccountValidationError(_('Informe um nome de conta com até 80 caracteres.'))
    if account_type not in ALLOWED_ACCOUNT_TYPES:
        raise AccountValidationError(_('Selecione um tipo de conta válido.'))
    if institution and len(institution) > 100:
        raise AccountValidationError(_('A instituição deve ter até 100 caracteres.'))
    if last_four and len(last_four) != 4:
        raise AccountValidationError(_('Informe exatamente os quatro últimos dígitos.'))
    if color not in ACCOUNT_COLOR_PALETTE:
        raise AccountValidationError(_('Selecione uma cor válida para a conta.'))

    return {
        'name': name,
        'account_type': account_type,
        'institution': institution or None,
        'last_four': last_four or None,
        'color': color,
        'initial_balance': initial_balance,
    }


def get_account_summaries(user_id: int) -> tuple[list[dict[str, Any]], Decimal]:
    accounts = FinancialAccount.query.filter_by(user_id=user_id).order_by(
        FinancialAccount.is_active.desc(),
        FinancialAccount.name.asc(),
    ).all()
    if not accounts:
        return [], Decimal('0.00')

    finance_rows = db.session.query(
        Finance.account_id,
        func.sum(
            case(
                (Finance.type == 'Receita', Finance.value),
                else_=-Finance.value,
            )
        ),
    ).filter(
        Finance.user_id == user_id,
        Finance.account_id.isnot(None),
        Finance.status == 'Pago',
    ).group_by(Finance.account_id).all()
    finance_totals = {row[0]: Decimal(str(row[1] or 0)) for row in finance_rows}

    outgoing_rows = db.session.query(
        AccountTransfer.source_account_id,
        func.sum(AccountTransfer.amount),
    ).filter(AccountTransfer.user_id == user_id).group_by(AccountTransfer.source_account_id).all()
    outgoing = {row[0]: Decimal(str(row[1] or 0)) for row in outgoing_rows}

    incoming_rows = db.session.query(
        AccountTransfer.destination_account_id,
        func.sum(AccountTransfer.amount),
    ).filter(AccountTransfer.user_id == user_id).group_by(AccountTransfer.destination_account_id).all()
    incoming = {row[0]: Decimal(str(row[1] or 0)) for row in incoming_rows}

    unreconciled_rows = db.session.query(
        BankTransaction.account_id,
        func.count(BankTransaction.id),
    ).filter(
        BankTransaction.user_id == user_id,
        BankTransaction.reconciled_finance_id.is_(None),
    ).group_by(BankTransaction.account_id).all()
    unreconciled = {row[0]: int(row[1] or 0) for row in unreconciled_rows}

    summaries = []
    consolidated = Decimal('0.00')
    for account in accounts:
        balance = (
            Decimal(str(account.initial_balance or 0))
            + finance_totals.get(account.id, Decimal('0'))
            + incoming.get(account.id, Decimal('0'))
            - outgoing.get(account.id, Decimal('0'))
        ).quantize(Decimal('0.01'))
        if account.is_active:
            consolidated += balance
        summaries.append(
            {
                'account': account,
                'balance': balance,
                'type_label': get_account_type_label(account.account_type),
                'unreconciled_count': unreconciled.get(account.id, 0),
            }
        )
    return summaries, consolidated.quantize(Decimal('0.01'))


def create_transfer(user_id: int, data) -> AccountTransfer:
    source_id = data.get('source_account_id', type=int)
    destination_id = data.get('destination_account_id', type=int)
    if not source_id or not destination_id or source_id == destination_id:
        raise AccountValidationError(_('Selecione contas de origem e destino diferentes.'))

    accounts = FinancialAccount.query.filter(
        FinancialAccount.user_id == user_id,
        FinancialAccount.id.in_([source_id, destination_id]),
        FinancialAccount.is_active.is_(True),
    ).all()
    if len(accounts) != 2:
        raise AccountValidationError(_('Uma das contas selecionadas não está disponível.'))

    amount = parse_decimal(data.get('amount'))
    if amount <= 0:
        raise AccountValidationError(_('O valor da transferência deve ser maior que zero.'))
    try:
        transfer_date = datetime.strptime(data.get('transfer_date') or '', '%Y-%m-%d').date()
    except ValueError as exc:
        raise AccountValidationError(_('Informe uma data válida para a transferência.')) from exc

    description = (data.get('description') or '').strip()[:140] or None
    return AccountTransfer(
        user_id=user_id,
        source_account_id=source_id,
        destination_account_id=destination_id,
        amount=amount,
        transfer_date=transfer_date,
        description=description,
    )


def validate_import_profile_payload(data) -> dict[str, Any]:
    name = (data.get('name') or '').strip()
    file_type = (data.get('file_type') or '').strip().lower()
    if not name or len(name) > 80:
        raise AccountValidationError(_('Informe um nome de perfil com até 80 caracteres.'))
    if file_type not in ALLOWED_PROFILE_FILE_TYPES:
        raise AccountValidationError(_('O perfil deve usar CSV ou XLSX.'))

    mapping = {
        'date': (data.get('date_column') or '').strip(),
        'description': (data.get('description_column') or '').strip(),
        'amount': (data.get('amount_column') or '').strip(),
        'reference': (data.get('reference_column') or '').strip(),
    }
    if not mapping['date'] or not mapping['description'] or not mapping['amount']:
        raise AccountValidationError(_('Mapeie as colunas de data, descrição e valor.'))
    delimiter = (data.get('delimiter') or ';').strip()
    if delimiter == r'\t':
        delimiter = '\t'
    if len(delimiter) != 1:
        raise AccountValidationError(_('Informe um delimitador CSV válido.'))
    return {
        'name': name,
        'file_type': file_type,
        'delimiter': delimiter if file_type == 'csv' else None,
        'mapping_json': json.dumps(mapping, ensure_ascii=False),
    }


def import_bank_statement(
    uploaded_file: FileStorage,
    account: FinancialAccount,
    profile: BankImportProfile | None = None,
    max_rows: int = MAX_STATEMENT_ROWS,
    max_file_size: int = MAX_STATEMENT_SIZE,
) -> StatementImportResult:
    filename = (uploaded_file.filename or '').strip()
    extension = os.path.splitext(filename)[1].lower()
    uploaded_file.stream.seek(0, os.SEEK_END)
    file_size = uploaded_file.stream.tell()
    uploaded_file.stream.seek(0)
    if not filename or file_size <= 0:
        raise AccountValidationError(_('Selecione um extrato não vazio para importar.'))
    if file_size > max_file_size:
        raise AccountValidationError(_('O extrato excede o limite de 10 MB.'))

    if extension == '.ofx':
        rows = _read_ofx_rows(uploaded_file)
        source = 'ofx'
    elif extension in {'.csv', '.xlsx'}:
        if profile is None or profile.file_type != extension.lstrip('.'):
            raise AccountValidationError(_('Selecione um perfil compatível com o arquivo enviado.'))
        rows = _read_profile_rows(uploaded_file, profile)
        source = profile.file_type
    else:
        raise AccountValidationError(_('Formato inválido. Utilize OFX, CSV ou XLSX.'))

    existing = {
        value
        for (value,) in db.session.query(BankTransaction.fingerprint).filter_by(
            account_id=account.id
        ).all()
    }
    transactions = []
    duplicate_rows = 0
    for index, row in enumerate(rows, start=1):
        if index > max_rows:
            raise AccountValidationError(
                _('O extrato excede o limite de %(count)d movimentações.', count=max_rows)
            )
        fingerprint = _transaction_fingerprint(account.id, row)
        if fingerprint in existing:
            duplicate_rows += 1
            continue
        existing.add(fingerprint)
        transactions.append(
            BankTransaction(
                user_id=account.user_id,
                account_id=account.id,
                external_id=(row.get('reference') or '')[:140] or None,
                fingerprint=fingerprint,
                transaction_date=row['date'],
                description=(row.get('description') or 'Movimentação importada')[:255],
                amount=row['amount'],
                source=source,
            )
        )
    if not transactions and not duplicate_rows:
        raise AccountValidationError(_('Nenhuma movimentação válida foi encontrada no extrato.'))
    return StatementImportResult(len(transactions), duplicate_rows, transactions)


def _read_ofx_rows(uploaded_file: FileStorage) -> list[dict[str, Any]]:
    raw = uploaded_file.stream.read()
    uploaded_file.stream.seek(0)
    text = None
    for encoding in ('utf-8-sig', 'cp1252', 'latin-1'):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        raise AccountValidationError(_('Não foi possível ler o arquivo OFX.'))

    blocks = re.findall(
        r'<STMTTRN>(.*?)(?=<STMTTRN>|</BANKTRANLIST>|</STMTTRN>)',
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    rows = []
    for block in blocks:
        amount_text = _ofx_tag(block, 'TRNAMT')
        date_text = _ofx_tag(block, 'DTPOSTED')
        if not amount_text or not date_text:
            continue
        try:
            transaction_date = datetime.strptime(date_text[:8], '%Y%m%d').date()
            amount = parse_decimal(amount_text)
        except (ValueError, AccountValidationError):
            continue
        rows.append(
            {
                'date': transaction_date,
                'description': _ofx_tag(block, 'MEMO') or _ofx_tag(block, 'NAME') or 'Movimentação OFX',
                'amount': amount,
                'reference': _ofx_tag(block, 'FITID'),
            }
        )
    return rows


def _ofx_tag(block: str, tag: str) -> str:
    match = re.search(
        rf'<{tag}>\s*([^<\r\n]+)',
        block,
        flags=re.IGNORECASE,
    )
    return match.group(1).strip() if match else ''


def _read_profile_rows(
    uploaded_file: FileStorage,
    profile: BankImportProfile,
) -> list[dict[str, Any]]:
    mapping = json.loads(profile.mapping_json)
    if profile.file_type == 'csv':
        raw = uploaded_file.stream.read()
        uploaded_file.stream.seek(0)
        text = None
        for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise AccountValidationError(_('Não foi possível decodificar o arquivo CSV.'))
        source_rows = list(csv.DictReader(io.StringIO(text), delimiter=profile.delimiter or ';'))
    else:
        uploaded_file.stream.seek(0)
        workbook = load_workbook(uploaded_file.stream, read_only=True, data_only=True)
        try:
            iterator = workbook.active.iter_rows(values_only=True)
            header = next(iterator, None)
            if not header:
                raise AccountValidationError(_('A planilha está vazia.'))
            headers = [str(value or '').strip() for value in header]
            source_rows = [
                {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
                for values in iterator
            ]
        finally:
            workbook.close()

    rows = []
    for raw_row in source_rows:
        if not any(str(value or '').strip() for value in raw_row.values()):
            continue
        try:
            rows.append(
                {
                    'date': _parse_statement_date(raw_row.get(mapping['date'])),
                    'description': str(raw_row.get(mapping['description']) or '').strip(),
                    'amount': parse_decimal(raw_row.get(mapping['amount'])),
                    'reference': str(raw_row.get(mapping.get('reference')) or '').strip(),
                }
            )
        except (AccountValidationError, ValueError):
            continue
    return rows


def _parse_statement_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(value))
    text = str(value or '').strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise AccountValidationError(_('Data inválida no extrato.'))


def _transaction_fingerprint(account_id: int, row: dict[str, Any]) -> str:
    identity = '|'.join(
        [
            str(account_id),
            row['date'].isoformat(),
            str(row['amount']),
            str(row.get('reference') or '').strip().lower(),
            str(row.get('description') or '').strip().lower(),
        ]
    )
    return hashlib.sha256(identity.encode('utf-8')).hexdigest()


def get_reconciliation_candidates(
    user_id: int,
    transaction: BankTransaction,
    limit: int = 8,
) -> list[Finance]:
    expected_type = 'Receita' if Decimal(str(transaction.amount)) > 0 else 'Despesa'
    expected_value = abs(Decimal(str(transaction.amount)))
    date_start = transaction.transaction_date - timedelta(days=7)
    date_end = transaction.transaction_date + timedelta(days=7)
    candidates = Finance.query.filter(
        Finance.user_id == user_id,
        Finance.type == expected_type,
        Finance.status == 'Pago',
        Finance.value.between(float(expected_value - Decimal('0.005')), float(expected_value + Decimal('0.005'))),
        Finance.due_date.between(date_start, date_end),
    ).order_by(Finance.due_date.asc(), Finance.id.desc()).all()
    # Keep this ordering portable between SQLite (desktop) and MySQL (server).
    # SQLite's julianday() previously made the reconciliation screen fail online.
    candidates.sort(
        key=lambda finance: (
            abs((finance.due_date - transaction.transaction_date).days),
            -finance.id,
        )
    )
    return candidates[:limit]


def reconcile_transaction(
    transaction: BankTransaction,
    finance: Finance,
) -> None:
    if transaction.user_id != finance.user_id:
        raise AccountValidationError(_('O lançamento selecionado não pertence a esta conta.'))
    expected_type = 'Receita' if Decimal(str(transaction.amount)) > 0 else 'Despesa'
    if finance.type != expected_type or abs(Decimal(str(transaction.amount))) != Decimal(str(finance.value)).quantize(Decimal('0.01')):
        raise AccountValidationError(_('Tipo ou valor incompatível com a movimentação bancária.'))
    transaction.reconciled_finance_id = finance.id
    transaction.reconciled_at = utcnow_naive()
    if finance.account_id is None:
        finance.account_id = transaction.account_id


def create_finance_from_transaction(transaction: BankTransaction) -> Finance:
    entry_type = 'Receita' if Decimal(str(transaction.amount)) > 0 else 'Despesa'
    finance = Finance(
        description=transaction.description[:100],
        value=float(abs(Decimal(str(transaction.amount)))),
        category='Outros',
        subcategory='Outras receitas' if entry_type == 'Receita' else 'Outras despesas',
        type=entry_type,
        status='Pago',
        due_date=transaction.transaction_date,
        payment_date=transaction.transaction_date,
        payment_method='Transferência / PIX',
        observations='Criado durante a conciliação bancária.',
        user_id=transaction.user_id,
        account_id=transaction.account_id,
    )
    db.session.add(finance)
    db.session.flush()
    reconcile_transaction(transaction, finance)
    return finance
